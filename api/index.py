import os
import re
import json
import google.generativeai as genai
from flask_cors import CORS
from PyPDF2 import PdfReader
from flask import Flask, render_template, request, jsonify
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_google_genai import GoogleGenerativeAIEmbeddings
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_community.vectorstores import FAISS
from langchain.chains.question_answering import load_qa_chain
from langchain.prompts import PromptTemplate
from dotenv import load_dotenv
from docx import Document
import csv
from openpyxl import load_workbook
import logging

app = Flask(__name__)
CORS(app)

# Configure logging
logging.basicConfig(level=logging.INFO)

load_dotenv()
api = os.getenv("GOOGLE_API_KEY")
genai.configure(api_key=api)

def extract_text_from_docx(docx_file):
    try:
        document = Document(docx_file)
        text = '\n'.join([paragraph.text for paragraph in document.paragraphs])
        logging.info(f"Extracted text from DOCX file: {docx_file.filename}")
        return text
    except Exception as e:
        logging.error(f"Error extracting text from DOCX file: {e}")
        return ""

def extract_text_from_xlsx(xlsx_file):
    try:
        workbook = load_workbook(xlsx_file.stream, data_only=True)
        text = ""
        for sheet in workbook:
            for row in sheet.iter_rows(values_only=True):
                row_text = ", ".join([str(cell) for cell in row if cell is not None])
                text += row_text + "\n"
        logging.info(f"Extracted text from XLSX file: {xlsx_file.filename}")
        return text
    except Exception as e:
        logging.error(f"Error extracting text from XLSX file: {e}")
        return ""

def extract_text_from_csv(csv_file):
    try:
        content = []
        stream = csv_file.stream.read().decode('utf-8').splitlines()
        reader = csv.reader(stream)
        for row in reader:
            content.append(', '.join(row))
        logging.info(f"Extracted text from CSV file: {csv_file.filename}")
        return '\n'.join(content)
    except Exception as e:
        logging.error(f"Error extracting text from CSV file: {e}")
        return ""

def get_text_from_documents(docs):
    file_contents = {}
    for doc in docs:
        try:
            if doc.filename.endswith('.pdf'):
                pdf_reader = PdfReader(doc.stream)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                file_contents[doc.filename] = text
                logging.info(f"Extracted text from PDF file: {doc.filename}")
            elif doc.filename.endswith('.docx'):
                file_contents[doc.filename] = extract_text_from_docx(doc)
            elif doc.filename.endswith('.csv'):
                file_contents[doc.filename] = extract_text_from_csv(doc)
            elif doc.filename.endswith('.xlsx'):
                file_contents[doc.filename] = extract_text_from_xlsx(doc)
        except Exception as e:
            logging.error(f"Error extracting text from file {doc.filename}: {e}")
    return file_contents

def get_text_chunks(text):
    try:
        text_splitter = RecursiveCharacterTextSplitter(chunk_size=10000, chunk_overlap=1000)
        chunks = text_splitter.split_text(text)
        logging.info("Text split into chunks successfully")
        return chunks
    except Exception as e:
        logging.error(f"Error splitting text into chunks: {e}")
        return []

def get_vectors(text_chunks):
    try:
        embeddings = GoogleGenerativeAIEmbeddings(model='models/embedding-001')
        vector_store = FAISS.from_texts(text_chunks, embedding=embeddings)
        vector_store.save_local("faiss_index")
        logging.info("Vectors generated and saved successfully")
    except Exception as e:
        logging.error(f"Error generating vectors: {e}")

def convo_chain():
    prompt_template = """
       You are an expert in writing academic research papers in the IEEE format. Create a research paper using the provided text. Follow these detailed instructions:

        1. **Abstract**: Summarize the main objectives, methods, results, and conclusions of the research in approximately 150 words.

        2. **Introduction**: Introduce the topic, outline the problem, and state the objectives of the research in more than 300 words.

        3. **Literature Review**: Discuss previous research on the topic, highlight gaps, and explain how this research fills those gaps.

        4. **Methodology**: Describe the methods and procedures used in the research in detail.

        5. **Results and Discussion**: Present the research findings, interpret the results, and discuss their implications.

        6. **Conclusion**: Summarize the key findings, their significance, and suggest areas for future research.

        7. **References**: List all references in the following format: [Number] Author's Initial(s). Author's Last Name, "Title of paper," Title of Journal, vol. number, no. number, pp. page range, Month Year. DOI.

        Ensure the paper is structured logically and adheres to the IEEE formatting guidelines. Only include references from the provided text.
        Text: {context}
        
    """
    prompt = PromptTemplate(template=prompt_template, input_variables=['context'])
    model = ChatGoogleGenerativeAI(model='gemini-pro', temperature=0.3)
    chain = load_qa_chain(model, chain_type='stuff', prompt=prompt)
    return chain

def user_input(user_question):
    try:
        embeddings = GoogleGenerativeAIEmbeddings(model='models/embedding-001')
        new_db = FAISS.load_local("faiss_index", embeddings, allow_dangerous_deserialization=True)
        docs = new_db.similarity_search(user_question)
        chain = convo_chain()
        response = chain({"input_documents": docs, "question": user_question}, return_only_outputs=True)
        logging.info("User input processed successfully")
        return response['output_text']
    except Exception as e:
        logging.error(f"Error processing user input: {e}")
        return str(e)

def clean_for_insights(qa_results):
    try:
        cleaned_text = re.sub(r'\|', '', qa_results)
        cleaned_text = re.sub(r'\s*\n\s*', ' ', cleaned_text).strip()
        cleaned_text = re.sub(r'\s{2,}', ' ', cleaned_text)
        logging.info("Text cleaned for insights successfully")
        return cleaned_text
    except Exception as e:
        logging.error(f"Error cleaning text for insights: {e}")
        return qa_results

def generate_insights(llm_response):
    try:
        overall_insight_question = "Add Future scope what could be done in it?"
        overall_insight = user_input(overall_insight_question + llm_response)
        cleaned_text_for_insights = clean_for_insights(overall_insight)
        logging.info("Insights generated successfully")
        return {"Overall Insight": cleaned_text_for_insights}
    except Exception as e:
        logging.error(f"Error generating insights: {e}")
        return {"Overall Insight": ""}

@app.route('/')
def index():
    return render_template('index.html')

def format_references(text):
    try:
        references = re.findall(r'\[(\d+)\] (.+?)(?=\[\d+\]|\Z)', text, re.DOTALL)
        formatted_references = "\n".join([f"[{num}] {ref.strip()}" for num, ref in references])
        logging.info("References formatted successfully")
        return formatted_references
    except Exception as e:
        logging.error(f"Error formatting references: {e}")
        return text

@app.route('/api/process', methods=['POST'])
def process_files():
    try:
        docs = request.files.getlist('docs')
        file_contents = get_text_from_documents(docs)
        combined_text = "\n".join(file_contents.values())
        text_chunks = get_text_chunks(combined_text)
        get_vectors(text_chunks)
        
        qa_results = user_input(combined_text)
        formatted_references = format_references(qa_results)
        
        cleaned_text_for_insights = clean_for_insights(qa_results)
        insights = generate_insights(cleaned_text_for_insights)

        logging.info("Files processed successfully")
        return jsonify({"insights": insights, "references": formatted_references})
    except Exception as e:
        logging.error(f"Error processing files: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(host="127.0.0.1",port="8000")